import pandas as pd
from openpyxl.utils.cell import range_boundaries

def paste_values(pyxl, df):
    """ Paste values from a DataFrame column into an openpyxl workbook"""
    
    sheet_names = pyxl.sheetnames
    merged_loc = pd.DataFrame()

    for sheet in sheet_names:
        merged = list(pyxl[sheet].merged_cells.ranges)
        merged_list = []
        for i in range(len(merged)):
            str(merged[i])
            merged_list.append(range_boundaries(str(merged[i]))[:2])
        merged_loc = pd.concat([merged_loc, pd.DataFrame({sheet: merged_list})], axis=1)

    for i in range(len(df)):

        sheet = pyxl[df["sheets"][i]]

        if len(df["cell_shapes"][i]) == 2:

            if df["cell_values"][i] is not None:
                sheet[df["cell_ranges"][i]].value = df["cell_values"][i][0][0]
        else:

            tuple_to_check = (df["cell_shapes"][i][0], df["cell_shapes"][i][1]) 
            print(f"Values: {df['cell_ranges'][i]} in sheet: {df['sheets'][i]}")

            if df["cell_values"][i] is not None: # avoiding formulas

                if tuple_to_check in merged_loc[df["sheets"][i]].values.tolist(): # merged cells check

                    print(f"Merged detected: {tuple_to_check}; sheet: {df['sheets'][i]}")
                    df["cell_values"][i] = [df["cell_values"][i][0]]

                    if len(df["cell_shapes"][i]) == (df["cell_shapes"][i][3]+1 - df["cell_shapes"][i][1]): # enlarged ranges check
                        for row in range(df["cell_shapes"][i][1], df["cell_shapes"][i][3] + 1):
                            sheet.cell(row=row, column=df["cell_shapes"][i][0]).value = df["cell_values"][i][row - df["cell_shapes"][i][1]][0]
                    else:
                        for j in range(df["cell_shapes"][i][3]+1 - df["cell_shapes"][i][1] - len(df["cell_values"][i])):
                            df["cell_values"][i].append([None])
                        for row in range(df["cell_shapes"][i][1], df["cell_shapes"][i][3] + 1):
                            sheet.cell(row=row, column=df["cell_shapes"][i][0]).value = df["cell_values"][i][row - df["cell_shapes"][i][1]][0]

                else:

                    if len(df["cell_shapes"][i]) == (df["cell_shapes"][i][3]+1 - df["cell_shapes"][i][1]):
                        for row in range(df["cell_shapes"][i][1], df["cell_shapes"][i][3] + 1):
                            for col in range(df["cell_shapes"][i][0], df["cell_shapes"][i][2] + 1):
                                sheet.cell(row=row, column=col).value = df["cell_values"][i][row - df["cell_shapes"][i][1]][col - df["cell_shapes"][i][0]]
                    else:
                        for j in range(df["cell_shapes"][i][3]+1 - df["cell_shapes"][i][1] - len(df["cell_values"][i])):
                            df["cell_values"][i].append([None])
                        for row in range(df["cell_shapes"][i][1], df["cell_shapes"][i][3] + 1):
                            for col in range(df["cell_shapes"][i][0], df["cell_shapes"][i][2] + 1):
                                sheet.cell(row=row, column=col).value = df["cell_values"][i][row - df["cell_shapes"][i][1]][col - df["cell_shapes"][i][0]]