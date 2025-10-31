from openpyxl import load_workbook
import pandas as pd

from outils.access_missingNR_table import access_missingNR_table
from outils.apply_changes_NR import apply_changes_NR
from outils.clean_NR_with_no_data import clean_NR_with_no_data
from outils.paste_values import paste_values
from outils.access_NR_table import access_NR_table
from outils.copy_values import copy_values

old_wb = load_workbook("Template/VSME-Digital-Template-Sample-1.0.0.xlsx")
new_wb_empty = load_workbook("Template/VSME-Digital-Template-1.1.1.xlsx")

version_cell = old_wb["Introduction"].cell(row=1, column=3).value
version_cell_new = "1.1.1" # new_wb_empty["Introduction"].cell(row=1, column=3).value TO APPLY AFTER SERGIO'S CHANGES

df_old = access_NR_table(old_wb.defined_names)
df_new = access_NR_table(new_wb_empty.defined_names)

missingNR_df = pd.read_pickle("pickles/missingNR_df.pkl") # pickle with missing name ranges across versions

missingNR_df_old = access_missingNR_table(missingNR_df, version_cell)
missingNR_df_new = access_missingNR_table(missingNR_df, version_cell_new)

df_old_wv = copy_values(old_wb, df_old, key="name_ranges")
missingNR_df_old_values = copy_values(old_wb, missingNR_df_old, key=None)

df_old_tomerge = df_old_wv[["name_ranges", "cell_values"]]
df_old_tomerge = apply_changes_NR(df_old_tomerge, version_cell, version_cell_new) # apply migration NR changes
df_old_tomerge = clean_NR_with_no_data(df_old_tomerge) # clean NRs which have no data to transfer

df_new_wv = df_new.merge(df_old_tomerge) 
missingNR_df_new_wv = pd.concat([missingNR_df_new, missingNR_df_old_values], axis=1)

paste_values(new_wb_empty, missingNR_df_new_wv)
paste_values(new_wb_empty, df_new_wv, NR=True)

new_wb_empty.save("Template/try100.xlsx")

# anti_join = df_old[~df_old['name_ranges'].isin(df_new['name_ranges'])]
# print(anti_join)

# df_new_wv.dropna()

# df_new_wv.to_pickle("df1.pkl")